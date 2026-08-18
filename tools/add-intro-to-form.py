#!/usr/bin/env python3
"""기존 결과문구 넘버스 양식에 'F_한줄소개' 시트를 덧붙인다 (기존 시트·내용 보존).

사용:  python3 tools/add-intro-to-form.py [입력.numbers] [출력.numbers]
기본:  입력=docs/용신_결과문구_작성양식.numbers, 출력=입력과 동일(제자리 갱신)

한 줄 소개 문장 = 템플릿 + 계절(월지 12) + 색동물(일주 60갑자).
색·동물을 조합 자동생성하지 않고 60갑자 전부를 기획이 직접 지정하는 방식(기획 결정).
기본값은 색(일간 오행)×동물(일지)로 미리 채워두므로, 마음에 안 드는 칸만 고치면 된다.

동작: 넘버스 → xlsx(Numbers 앱) → openpyxl로 시트 추가 → 다시 넘버스(Numbers 앱). macOS 전용.
"""
import sys
import platform
import subprocess
import tempfile
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent
DEFAULT = ROOT / "docs" / "용신_결과문구_작성양식.numbers"

SHEET = "F_한줄소개"
TEMPLATE_DEFAULT = "당신은 {계절}에 태어난 작고 소중한 {색동물}"

GAN = "甲乙丙丁戊己庚辛壬癸"
JI = "子丑寅卯辰巳午未申酉戌亥"
GAN_KO = "갑을병정무기경신임계"
JI_KO = "자축인묘진사오미신유술해"
# 천간 오행(목화토금수) 인덱스 → 색 표현
STEM_WX = [0, 0, 1, 1, 2, 2, 3, 3, 4, 4]
COLOR = ["푸른", "붉은", "누런", "하얀", "검은"]
ANIMAL = ["쥐", "소", "호랑이", "토끼", "용", "뱀", "말", "양", "원숭이", "닭", "개", "돼지"]
# 월지(지지 인덱스 0=子 … 11=亥) → 계절 기본 표현
SEASON = ["한겨울", "한겨울", "이른 봄", "봄", "늦봄", "이른 여름",
          "한여름", "늦여름", "이른 가을", "가을", "늦가을", "초겨울"]

TITLE_F = Font(name="맑은 고딕", size=15, bold=True, color="1D1813")
H1 = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
BOLD = Font(name="맑은 고딕", size=10, bold=True)
BODY = Font(name="맑은 고딕", size=10)
MUTED = Font(name="맑은 고딕", size=9, color="6B6255")
DRAFT = Font(name="맑은 고딕", size=10, color="8A6F45", italic=True)
HEAD_FILL = PatternFill("solid", fgColor="4A3823")
INPUT_FILL = PatternFill("solid", fgColor="FFF9DB")   # 노란색 = 기획이 채울 칸
WRAP = Alignment(vertical="center", wrap_text=True)


def numbers_to_xlsx(src, dst):
    _osa(src, dst, export=True)


def xlsx_to_numbers(src, dst):
    if dst.exists():
        dst.unlink()
    _osa(src, dst, export=False)


def _osa(src, dst, export):
    if platform.system() != "Darwin":
        raise SystemExit("❌ 이 작업은 macOS의 Numbers 앱이 필요합니다.")
    verb = ("export d to dst as Microsoft Excel" if export else "save d in dst")
    script = ('on run argv\n'
              '  set src to POSIX file (item 1 of argv)\n'
              '  set dst to POSIX file (item 2 of argv)\n'
              '  tell application "Numbers"\n'
              '    set d to open src\n'
              '    delay 1\n'
              f'    {verb}\n'
              '    close d saving no\n'
              '  end tell\n'
              'end run\n')
    r = subprocess.run(["osascript", "-", str(src), str(dst)],
                       input=script, capture_output=True, text=True)
    if r.returncode != 0 or not Path(dst).exists():
        raise SystemExit("❌ Numbers 변환 실패: " + (r.stderr.strip() or "알 수 없는 오류"))


def add_intro_sheet(wb):
    if SHEET in wb.sheetnames:
        print(f"ℹ️  '{SHEET}' 시트가 이미 있어 건너뜁니다.")
        return False
    # 조합미리보기 앞(있으면)에 끼워넣어 작성 시트끼리 모이게
    idx = wb.sheetnames.index("조합미리보기") if "조합미리보기" in wb.sheetnames else len(wb.sheetnames)
    ws = wb.create_sheet(SHEET, idx)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 34

    def head(r, a, b, c, d):
        for col, val in zip("ABCD", (a, b, c, d)):
            cell = ws[f"{col}{r}"]
            cell.value = val
            cell.font = H1
            cell.fill = HEAD_FILL
            cell.alignment = WRAP

    def row(r, key, label, sample):
        ws[f"A{r}"] = key; ws[f"A{r}"].font = MUTED
        ws[f"B{r}"] = label; ws[f"B{r}"].font = BODY
        ws[f"C{r}"] = sample; ws[f"C{r}"].font = DRAFT
        c = ws[f"D{r}"]; c.value = sample; c.font = BODY; c.fill = INPUT_FILL; c.alignment = WRAP

    ws["A1"] = "F. 한 줄 소개"; ws["A1"].font = TITLE_F
    ws["A2"] = ("결과 화면 맨 위에 붙는 요약 문장입니다.  "
                "예) 당신은 한겨울에 태어난 작고 소중한 하얀쥐")
    ws["A2"].font = MUTED
    ws["A3"] = ("계절은 월지(태어난 달)에서, 색동물은 일주(태어난 날의 60갑자)에서 자동으로 골라 끼웁니다. "
                "노란 칸만 원하는 표현으로 고치면 됩니다.")
    ws["A3"].font = MUTED

    # ① 템플릿
    ws["A5"] = "① 문장 틀"; ws["A5"].font = BOLD
    head(6, "키(코드)", "설명", "예시", "✏️ 문구 (여기에 작성)")
    row(7, "template", "문장 틀", TEMPLATE_DEFAULT)
    ws["A8"] = "※ {계절}·{색동물}은 아래 표에서 자동으로 치환됩니다. 문장 틀만 바꾸고 싶으면 D7만 고치세요."
    ws["A8"].font = MUTED

    # ② 계절 (월지 12)
    ws["A10"] = "② 계절 — 월지(태어난 달의 지지) 12"; ws["A10"].font = BOLD
    head(11, "키(코드)", "월지", "예시", "✏️ 계절 표현")
    for i in range(12):
        row(12 + i, f"s{i}", f"{JI[i]}({JI_KO[i]})월", SEASON[i])

    # ③ 색동물 (일주 60갑자)
    base = 12 + 12 + 1  # = 25
    ws[f"A{base}"] = "③ 색동물 — 일주(태어난 날의 60갑자) 60"; ws[f"A{base}"].font = BOLD
    head(base + 1, "키(코드)", "일주(갑자)", "예시", "✏️ 색·동물 표현")
    start = base + 2  # = 27
    for n in range(60):
        stem, branch = n % 10, n % 12
        default = COLOR[STEM_WX[stem]] + ANIMAL[branch]
        label = f"{GAN[stem]}{JI[branch]}({GAN_KO[stem]}{JI_KO[branch]})"
        row(start + n, f"g{n}", label, default)
    return True


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    src = Path(args[0]).expanduser() if len(args) >= 1 else DEFAULT
    dst = Path(args[1]).expanduser() if len(args) >= 2 else src
    if not src.exists():
        raise SystemExit(f"❌ 입력 파일이 없습니다: {src}")

    tmp_in = Path(tempfile.mkdtemp()) / (src.stem + ".xlsx")
    print(f"→ 넘버스 → xlsx 변환 (Numbers 창이 잠깐 떴다 닫힙니다) …")
    numbers_to_xlsx(src, tmp_in)

    wb = load_workbook(tmp_in)
    added = add_intro_sheet(wb)
    if not added:
        return
    tmp_out = Path(tempfile.mkdtemp()) / (dst.stem + ".xlsx")
    wb.save(tmp_out)

    print(f"→ xlsx → 넘버스 변환 …")
    xlsx_to_numbers(tmp_out, dst)
    print(f"✅ '{SHEET}' 시트 추가 완료 → {dst}")
    print("   계절 12칸 · 색동물 60칸 (기본값 채워둠) · 문장 틀 1칸")


if __name__ == "__main__":
    main()
