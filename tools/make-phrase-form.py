#!/usr/bin/env python3
"""기획자용 결과 문구 작성 양식 생성 — 명세서 §11 블록 조립 구조.

사용:  python3 tools/make-phrase-form.py [출력경로]
기본:  policy/용신_결과문구_작성양식.numbers   (기획자가 맥 Numbers로 작업하므로)
       확장자를 .xlsx 로 주면 엑셀로 바로 생성한다.

.numbers 로 낼 때는 openpyxl 로 xlsx 를 만든 뒤 Numbers 앱을 시켜 변환한다(macOS 전용).
A·D·E 는 용신 오행 5종, B 는 §4 판정표 강약 5단계(engine/src/yongsin.js 의 LEVELS).
양식 구조를 바꾸면 tools/apply-phrases.py 의 파싱 위치도 함께 맞춰야 한다.
openpyxl 이 필요하다(생성 전용 도구라 개발자만 실행하면 됨).
"""
import sys
import platform
import subprocess
import tempfile
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

ROOT = Path(__file__).resolve().parent.parent
OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "policy" / "용신_결과문구_작성양식.numbers"


def xlsx_to_numbers(src, dst):
    """Numbers 앱을 시켜 xlsx → .numbers 로 변환. apply-phrases.py 의 역방향."""
    if platform.system() != "Darwin":
        raise RuntimeError(".numbers 생성은 macOS의 Numbers 앱이 필요합니다. "
                           "출력 경로를 .xlsx 로 주면 엑셀로 만듭니다.")
    if dst.exists():
        dst.unlink()
    script = (
        'on run argv\n'
        '  set src to POSIX file (item 1 of argv)\n'
        '  set dst to POSIX file (item 2 of argv)\n'
        '  tell application "Numbers"\n'
        '    set d to open src\n'
        '    delay 1\n'
        '    save d in dst\n'
        '    close d saving no\n'
        '  end tell\n'
        'end run\n')
    r = subprocess.run(["osascript", "-", str(src), str(dst)],
                       input=script, capture_output=True, text=True)
    if r.returncode != 0 or not dst.exists():
        raise RuntimeError("Numbers 변환 실패: " + (r.stderr.strip() or "알 수 없는 오류") +
                           "\n   Numbers 앱 설치 여부와 자동화 권한 허용을 확인하세요.")

FONT = "맑은 고딕"
TITLE = Font(name=FONT, size=15, bold=True, color="1D1813")
H1 = Font(name=FONT, size=11, bold=True, color="FFFFFF")
BOLD = Font(name=FONT, size=10, bold=True)
BODY = Font(name=FONT, size=10)
MUTED = Font(name=FONT, size=9, color="6B6255")
DRAFT = Font(name=FONT, size=10, color="8A6F45", italic=True)

HEAD_FILL = PatternFill("solid", fgColor="4A3823")
INPUT_FILL = PatternFill("solid", fgColor="FFF9DB")   # 노란색 = 기획이 채울 칸
FIXED_FILL = PatternFill("solid", fgColor="F2EDE0")

THIN = Side(style="thin", color="CBC1AA")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

# 오행: 한글 · 한자 · 구어명 · 희신 · 기신 · 색 · 방위 · 계절
WX = [("목", "木", "나무", "수", "금", "청(靑)", "동(東)", "봄"),
      ("화", "火", "불", "목", "수", "적(赤)", "남(南)", "여름"),
      ("토", "土", "땅", "화", "목", "황(黃)", "중앙", "환절기"),
      ("금", "金", "쇠", "토", "화", "백(白)", "서(西)", "가을"),
      ("수", "水", "물", "금", "토", "흑(黑)", "북(北)", "겨울")]

# §4 판정표 5단계 — 키 · 표시 · 언제 나오나
LEVELS = [
    ("strong", "신강", "아주 강할 때 · 신강 (총점 3 이상)"),
    ("midStrong", "중화(신강쪽)", "살짝 강할 때 · 중화(신강 쪽) (총점 1 ~ 2.9)"),
    ("neutral", "중화", "딱 균형일 때 · 중화 (총점 −0.9 ~ 0.9)"),
    ("midWeak", "중화(신약쪽)", "살짝 약할 때 · 중화(신약 쪽) (총점 −2.9 ~ −1)"),
    ("weak", "신약", "아주 약할 때 · 신약 (총점 −3 이하)"),
]

wb = Workbook()


def header(ws, row, cols):
    for i, (label, width) in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font, c.fill, c.alignment, c.border = H1, HEAD_FILL, CENTER, BOX
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def block_sheet(ws, title, desc, rows, count_note):
    """A~D 블록 공통: 키 | 언제 나오나 | 현재 draft | ✏️새 문구(입력) | 비고"""
    ws["A1"], ws["A1"].font = title, TITLE
    ws["A2"], ws["A2"].font = desc, MUTED
    ws["A3"] = count_note
    ws["A3"].font = Font(name=FONT, size=9, bold=True, color="A8352B")
    header(ws, 5, [("키(코드)", 13), ("언제 나오나", 36), ("현재 draft 문구 (개발용, 교체 대상)", 52),
                   ("✏️ 새 문구 (여기에 작성)", 52), ("비고", 20)])
    r = 6
    for key, cond, draft in rows:
        ws.cell(row=r, column=1, value=key).font = BOLD
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2, value=cond).font = BODY
        ws.cell(row=r, column=3, value=draft).font = DRAFT
        ws.cell(row=r, column=4).fill = INPUT_FILL
        ws.cell(row=r, column=4).font = BODY
        ws.cell(row=r, column=5).font = BODY
        for col in range(1, 6):
            cc = ws.cell(row=r, column=col)
            cc.border, cc.alignment = BOX, WRAP
            if col in (2, 3):
                cc.fill = FIXED_FILL
        ws.row_dimensions[r].height = 46
        r += 1


# ─────────────────────────── 작성안내
ws = wb.active
ws.title = "작성안내"
ws["A1"] = "부적집 · 용신 결과 문구 작성 양식"
ws["A1"].font = Font(name=FONT, size=17, bold=True, color="1D1813")
ws["A2"] = "결과 화면에 나가는 모든 문구를 여기서 채웁니다. 개발은 이 파일 그대로 반영합니다."
ws["A2"].font = MUTED
for col, w in [("A", 14), ("B", 26), ("C", 22), ("D", 20), ("E", 18), ("F", 14), ("G", 14), ("H", 12)]:
    ws.column_dimensions[col].width = w


def section(row, text, sub=None):
    ws.cell(row=row, column=1, value=text).font = Font(name=FONT, size=12, bold=True, color="8A6F45")
    if sub:
        ws.cell(row=row + 1, column=1, value=sub).font = MUTED
    return row + (2 if sub else 1)


r = section(4, "1. 채워야 할 것", "노란색 칸만 채우면 됩니다. 회색 칸은 참고용이니 건드리지 마세요.")
for i, h in enumerate(["시트", "개수", "내용"], start=1):
    c = ws.cell(row=r, column=i, value=h)
    c.font, c.fill, c.alignment, c.border = H1, HEAD_FILL, CENTER, BOX
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
r += 1
for name, cnt, desc in [
    ("A_용신선언", "5개", "용신 오행별로 \"필요한 기운은 ○예요\"를 선언하는 문장"),
    ("B_강약진단", "5개", "사주 기운 상태를 알려주는 첫 문장 (강약 5단계)"),
    ("C_조후보정", "2개", "사주가 춥거나 더울 때만 덧붙는 문장 (평범하면 생략됨)"),
    ("D_부적처방", "5개", "용신 오행별 부적 추천 멘트"),
    ("E_용신설명", "5개", "용신 오행별 상세 설명 + 키워드"),
]:
    ws.cell(row=r, column=1, value=name).font = BOLD
    ws.cell(row=r, column=2, value=cnt).font = BODY
    ws.cell(row=r, column=2).alignment = CENTER
    ws.cell(row=r, column=3, value=desc).font = BODY
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    for i in range(1, 9):
        ws.cell(row=r, column=i).border = BOX
    r += 1
ws.cell(row=r, column=1, value="합계 22개").font = Font(name=FONT, size=10, bold=True, color="A8352B")
r += 2

r = section(r, "2. 강약 5단계", "사주의 힘을 점수로 재서 다섯 구간으로 나눕니다. A·B 시트가 이 구간별로 갈립니다.")
for i, h in enumerate(["키(코드)", "판정", "점수 범위", "이 사람에게 필요한 것"], start=1):
    c = ws.cell(row=r, column=i, value=h)
    c.font, c.fill, c.alignment, c.border = H1, HEAD_FILL, CENTER, BOX
r += 1
for key, label, rng, need in [
    ("strong", "신강", "3 이상", "힘을 빼주는 기운"),
    ("midStrong", "중화(신강 쪽)", "1 ~ 2.9", "살짝 덜어주는 기운"),
    ("neutral", "중화", "−0.9 ~ 0.9", "부족한 곳을 채우는 기운"),
    ("midWeak", "중화(신약 쪽)", "−2.9 ~ −1", "살짝 보태주는 기운"),
    ("weak", "신약", "−3 이하", "북돋아 세워주는 기운"),
]:
    for i, v in enumerate([key, label, rng, need], start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font, c.border, c.fill = (BOLD if i == 1 else BODY), BOX, FIXED_FILL
        c.alignment = CENTER if i <= 3 else Alignment(vertical="center")
    r += 1
r += 1

r = section(r, "3. 문장 안에서 쓸 수 있는 변수", "아래 표기를 넣으면 사람마다 알맞은 오행 이름으로 자동 치환됩니다.")
for k, v in [("{용신}", "그 사람에게 필요한 기운 (예: 불)"),
             ("{희신}", "용신을 도와주는 기운 (예: 나무)"),
             ("{기신}", "용신을 방해하는 기운 (예: 물)")]:
    ws.cell(row=r, column=1, value=k).font = Font(name=FONT, size=10, bold=True, color="1F5C3A")
    ws.cell(row=r, column=2, value=v).font = BODY
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
ws.cell(row=r, column=1, value='예) "{기신}의 기운이 너무 강한 당신!"  →  "물의 기운이 너무 강한 당신!"').font = MUTED
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 2

r = section(r, "4. 문구가 조립되는 순서", "네 블록이 이 순서로 이어 붙어 결과 화면에 나갑니다.")
ws.cell(row=r, column=1, value="B (강약 진단)  →  A (용신 선언)  →  C (조후, 있을 때만)  →  D (부적 처방)").font = BOLD
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1
ws.cell(row=r, column=1, value="E (용신 설명)은 이 흐름과 별개로, 용신 한자 아래 영역에 들어갑니다.").font = MUTED
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 2

r = section(r, "5. 작성 시 지켜주세요", "명세서 §11-4 체크리스트")
for t in ["① 블록끼리 자연스럽게 이어지도록 문장 끝맺음을 통일해주세요. (A는 \"~예요.\", B는 \"~라서,\" 처럼)",
          "② C(조후)는 안 나오는 사람도 많습니다. C 없이 B→A→D만 읽어도 말이 되게 써주세요.",
          "③ A와 D는 둘 다 용신 기반이라 같은 말이 반복되기 쉽습니다. A는 진단, D는 처방으로 역할을 나눠주세요.",
          "④ 강약 5단계는 톤 차이입니다. 신강/신약은 단정적으로, 중화 3단계는 완만하게 쓰면 자연스럽습니다.",
          "⑤ 22개 전체가 하나의 관점(유파)으로 통일되어야 합니다.",
          "⑥ 조합미리보기 시트에서 75가지 실제 문장을 확인할 수 있습니다. 어색한 조합이 없는지 봐주세요."]:
    ws.cell(row=r, column=1, value=t).font = BODY
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
r += 1

r = section(r, "6. 오행 참조표", "용신이 정해지면 희신·기신은 자동으로 따라옵니다. (수정 불가)")
for i, h in enumerate(["오행", "한자", "구어 명칭", "희신(돕는 기운)", "기신(방해 기운)", "색", "방위", "계절"], start=1):
    c = ws.cell(row=r, column=i, value=h)
    c.font, c.fill, c.alignment, c.border = H1, HEAD_FILL, CENTER, BOX
r += 1
REF_START = r
for vals in WX:
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font, c.alignment, c.border, c.fill = BODY, CENTER, BOX, FIXED_FILL
    r += 1
REF_END = r - 1

# ─────────────────────────── A · B (5단계)
A_DRAFT = ["당신에게 지금 가장 필요한 기운은 '나무'예요.",
           "당신에게 지금 가장 필요한 기운은 '불'이에요.",
           "당신에게 지금 가장 필요한 기운은 '땅'이에요.",
           "당신에게 지금 가장 필요한 기운은 '쇠'예요.",
           "당신에게 지금 가장 필요한 기운은 '물'이에요."]
B_DRAFT = {
    "strong": "뭔가 부족하다기보다는 {기신}의 기운이 너무나 강한 당신!",
    "midStrong": "크게 치우치진 않았지만 {기신}의 기운이 살짝 도는 당신!",
    "neutral": "기운이 크게 치우치지 않아 팽팽하게 균형을 이루는 당신!",
    "midWeak": "크게 치우치진 않았지만 받쳐줄 힘이 조금 아쉬운 당신!",
    "weak": "기운이 여려서 든든하게 받쳐줄 힘이 필요한 당신!",
}
block_sheet(wb.create_sheet("A_용신선언"), "A. 용신 선언",
            "\"당신에게 필요한 기운은 ○예요\" — 용신을 선언하는 핵심 문장입니다. 용신 오행 5가지별로 씁니다.",
            [(f"{ko} ({han})", f"용신이 {spoken}일 때", A_DRAFT[i])
             for i, (ko, han, spoken, *_x) in enumerate(WX)], "채울 칸: 5개 (노란색)")
block_sheet(wb.create_sheet("B_강약진단"), "B. 강약 진단",
            "결과 화면 첫 문장입니다. 사주 기운이 어떤 상태인지 알려줍니다.",
            [(k, cond, B_DRAFT[k]) for k, _lab, cond in LEVELS], "채울 칸: 5개 (노란색)")
block_sheet(wb.create_sheet("C_조후보정"), "C. 조후 보정 (조건부)",
            "사주가 뚜렷하게 춥거나 더울 때만 덧붙습니다. 그렇지 않은 사람에겐 이 문장이 아예 안 나옵니다.",
            [("cold", "사주가 차가울 때 (한寒)", "게다가 사주가 차게 얼어 있어, 따뜻한 기운이 이 냉기를 녹여줘요."),
             ("hot", "사주가 뜨거울 때 (열熱)", "게다가 사주가 뜨겁게 달아 있어, 시원한 기운이 이 열기를 식혀줘요.")],
            "채울 칸: 2개 (노란색) · '평범(평)'인 사람은 이 블록이 생략되니 따로 쓸 필요 없습니다")
block_sheet(wb.create_sheet("D_부적처방"), "D. 부적 처방",
            "마지막 문장입니다. 용신 기운을 담은 부적을 권하는 멘트로, 용신 오행 5가지별로 씁니다.",
            [("목 (木)", "용신이 나무일 때", "쭉쭉 뻗는 나무의 힘으로 숨통을 틔워보는거 어때?"),
             ("화 (火)", "용신이 불일 때", "불의 힘으로 수마를 날려버리는거 어때?"),
             ("토 (土)", "용신이 땅일 때", "든든한 땅의 힘으로 중심을 잡아보는거 어때?"),
             ("금 (金)", "용신이 쇠일 때", "단단한 쇠의 힘으로 딱 끊어내는거 어때?"),
             ("수 (水)", "용신이 물일 때", "흐르는 물의 힘으로 열기를 식혀보는거 어때?")],
            "채울 칸: 5개 (노란색)")

# ─────────────────────────── E_용신설명
ws = wb.create_sheet("E_용신설명")
ws["A1"], ws["A1"].font = "E. 용신 오행별 설명", TITLE
ws["A2"] = "결과 화면에서 큰 용신 한자 아래에 들어갈 설명입니다. 용신 5가지에 대해 각각 작성해주세요."
ws["A2"].font = MUTED
ws["A3"] = "채울 칸: 5행 × 3열 = 15칸 (노란색). 색·방위·계절은 기본값을 넣어뒀으니 필요하면 고쳐주세요."
ws["A3"].font = Font(name=FONT, size=9, bold=True, color="A8352B")
header(ws, 5, [("용신", 9), ("한자", 8), ("구어 명칭", 11),
               ("✏️ 제목 (한 줄, 12자 내외)", 30), ("✏️ 설명 본문 (2~3문장)", 60),
               ("✏️ 키워드 (쉼표 구분, 3개)", 26), ("색", 11), ("방위", 10), ("계절", 11)])
EXAMPLE = {"화": ("밝히고 데우는 불",
                  "차게 식은 기운을 녹여 몸과 마음에 온기를 되돌려 줍니다. 표현이 살아나고 사람들 앞에 서는 일이 수월해집니다.",
                  "명예, 표현, 열정")}
r = 6
for ko, han, spoken, hee, gi, color, dirn, season in WX:
    ws.cell(row=r, column=1, value=ko).font = BOLD
    ws.cell(row=r, column=2, value=han).font = Font(name=FONT, size=12, bold=True, color="8A6F45")
    ws.cell(row=r, column=3, value=spoken).font = BODY
    ex = EXAMPLE.get(ko)
    for col, val in ((4, ex[0] if ex else None), (5, ex[1] if ex else None), (6, ex[2] if ex else None)):
        c = ws.cell(row=r, column=col, value=val)
        c.fill, c.font = INPUT_FILL, (DRAFT if ex else BODY)
    for col, val in ((7, color), (8, dirn), (9, season)):
        ws.cell(row=r, column=col, value=val).font = BODY
    for col in range(1, 10):
        c = ws.cell(row=r, column=col)
        c.border, c.alignment = BOX, (WRAP if col in (4, 5, 6) else CENTER)
        if col not in (4, 5, 6):
            c.fill = FIXED_FILL
    ws.row_dimensions[r].height = 62
    r += 1
ws.cell(row=6, column=4).comment = Comment(
    "화(火) 행은 작성 예시입니다. 형식만 참고하시고 내용은 새로 써주세요.", "개발", width=280, height=80)
ws.cell(row=r + 1, column=1, value="※ 화(火) 행에 채워둔 값은 형식 안내용 예시입니다. 실제 문구로 바꿔주세요.").font = \
    Font(name=FONT, size=9, color="A8352B")
ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=9)

# ─────────────────────────── F_한줄소개 (한 줄 소개 = 문장틀 + 계절 월지12 + 색동물 일주60갑자)
_GAN, _JI = "甲乙丙丁戊己庚辛壬癸", "子丑寅卯辰巳午未申酉戌亥"
_GAN_KO, _JI_KO = "갑을병정무기경신임계", "자축인묘진사오미신유술해"
_STEM_WX = [0, 0, 1, 1, 2, 2, 3, 3, 4, 4]
_COLOR = ["푸른", "붉은", "누런", "하얀", "검은"]
_ANIMAL = ["쥐", "소", "호랑이", "토끼", "용", "뱀", "말", "양", "원숭이", "닭", "개", "돼지"]
_SEASON = ["한겨울", "한겨울", "이른 봄", "봄", "늦봄", "이른 여름",
           "한여름", "늦여름", "이른 가을", "가을", "늦가을", "초겨울"]
_TEMPLATE = "당신은 {계절}에 태어난 작고 소중한 {색동물}"

ws = wb.create_sheet("F_한줄소개")
ws["A1"], ws["A1"].font = "F. 한 줄 소개", TITLE
ws["A2"] = "결과 화면 맨 위에 붙는 요약 문장입니다.  예) 당신은 한겨울에 태어난 작고 소중한 하얀쥐"
ws["A2"].font = MUTED
ws["A3"] = "계절은 월지(태어난 달), 색동물은 일주(태어난 날의 60갑자)에서 자동으로 끼웁니다. 노란 칸만 원하는 표현으로 고치세요."
ws["A3"].font = MUTED


def _f_row(r, key, label, sample):
    ws.cell(row=r, column=1, value=key).font = MUTED
    ws.cell(row=r, column=2, value=label).font = BODY
    ws.cell(row=r, column=3, value=sample).font = DRAFT
    c = ws.cell(row=r, column=4, value=sample)
    c.fill, c.font, c.alignment = INPUT_FILL, BODY, WRAP


ws["A5"], ws["A5"].font = "① 문장 틀", BOLD
header(ws, 6, [("키(코드)", 10), ("설명", 16), ("예시", 22), ("✏️ 문구", 34)])
_f_row(7, "template", "문장 틀", _TEMPLATE)
ws["A8"] = "※ {계절}·{색동물}은 아래 표에서 자동 치환됩니다. 문장 틀만 바꾸려면 D7만 고치세요."
ws["A8"].font = MUTED

ws["A10"], ws["A10"].font = "② 계절 — 월지(태어난 달의 지지) 12", BOLD
header(ws, 11, [("키(코드)", 10), ("월지", 16), ("예시", 22), ("✏️ 계절 표현", 34)])
for i in range(12):
    _f_row(12 + i, f"s{i}", f"{_JI[i]}({_JI_KO[i]})월", _SEASON[i])

_base = 25
ws[f"A{_base}"], ws[f"A{_base}"].font = "③ 색동물 — 일주(태어난 날의 60갑자) 60", BOLD
header(ws, _base + 1, [("키(코드)", 10), ("일주(갑자)", 16), ("예시", 22), ("✏️ 색·동물 표현", 34)])
for n in range(60):
    _st, _br = n % 10, n % 12
    _f_row(_base + 2 + n, f"g{n}", f"{_GAN[_st]}{_JI[_br]}({_GAN_KO[_st]}{_JI_KO[_br]})",
           _COLOR[_STEM_WX[_st]] + _ANIMAL[_br])

# ─────────────────────────── 조합미리보기 (5 오행 × 5 강약 × 3 조후 = 75)
ws = wb.create_sheet("조합미리보기")
ws["A1"], ws["A1"].font = "조합 미리보기 (75가지)", TITLE
ws["A2"] = "왼쪽 시트에 문구를 채우면 실제 손님이 보게 될 문장이 여기에 자동으로 조립됩니다. 어색한 조합이 없는지 확인해주세요."
ws["A2"].font = MUTED
ws["A3"] = "※ 변수({용신}·{기신}·{희신})는 각 행의 오행으로 치환해 보여줍니다. 이 시트는 확인용이라 채울 칸이 없습니다."
ws["A3"].font = Font(name=FONT, size=9, color="6B6255")
header(ws, 5, [("#", 5), ("용신", 9), ("강약", 15), ("조후", 10),
               ("용신", 9), ("희신", 9), ("기신", 9), ("실제로 나가는 문장", 120)])

CLIMS = [("한(寒)", "cold"), ("열(熱)", "hot"), ("평(平)", "")]
NAME_COL = f"작성안내!$C${REF_START}:$C${REF_END}"
WX_COL = f"작성안내!$A${REF_START}:$A${REF_END}"

r, n = 6, 1
for wx_i, (ko, han, spoken, hee, gi, *_x) in enumerate(WX):
    for lv_key, lv_label, _cond in LEVELS:
        for clim_ko, clim_key in CLIMS:
            ws.cell(row=r, column=1, value=n).alignment = CENTER
            ws.cell(row=r, column=2, value=f"{ko}({han})").alignment = CENTER
            ws.cell(row=r, column=3, value=lv_label).alignment = CENTER
            ws.cell(row=r, column=4, value=clim_ko).alignment = CENTER
            ws.cell(row=r, column=5, value=f"=INDEX({NAME_COL},{wx_i+1})").alignment = CENTER
            ws.cell(row=r, column=6, value=f'=INDEX({NAME_COL},MATCH("{hee}",{WX_COL},0))').alignment = CENTER
            ws.cell(row=r, column=7, value=f'=INDEX({NAME_COL},MATCH("{gi}",{WX_COL},0))').alignment = CENTER

            # T()로 감싸 빈 칸이 0으로 조립되는 것을 막는다 (INDEX는 빈 셀을 0으로 돌려줌)
            b = f'T(INDEX(B_강약진단!$D$6:$D$10,MATCH("{lv_key}",B_강약진단!$A$6:$A$10,0)))'
            a = f'T(INDEX(A_용신선언!$D$6:$D$10,{wx_i+1}))'
            d = f'T(INDEX(D_부적처방!$D$6:$D$10,{wx_i+1}))'
            if clim_key:
                c_ref = f'T(INDEX(C_조후보정!$D$6:$D$7,MATCH("{clim_key}",C_조후보정!$A$6:$A$7,0)))'
                raw = f'{b}&" "&{a}&" "&{c_ref}&" "&{d}'
            else:
                raw = f'{b}&" "&{a}&" "&{d}'
            sub = f'SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({raw},"{{용신}}",$E{r}),"{{희신}}",$F{r}),"{{기신}}",$G{r})'
            ws.cell(row=r, column=8, value=f'=IF(TRIM({sub})="","(아직 문구가 채워지지 않았습니다)",TRIM({sub}))')

            for col in range(1, 9):
                c = ws.cell(row=r, column=col)
                c.font, c.border = BODY, BOX
                if col == 8:
                    c.alignment = WRAP
                elif col in (5, 6, 7):
                    c.fill = FIXED_FILL
            ws.row_dimensions[r].height = 44
            r += 1
            n += 1

OUT.parent.mkdir(parents=True, exist_ok=True)
if OUT.suffix.lower() == ".numbers":
    tmp = Path(tempfile.mkdtemp()) / (OUT.stem + ".xlsx")
    wb.save(tmp)
    print("→ Numbers로 변환 중 (Numbers 창이 잠깐 떴다 닫힙니다) …")
    xlsx_to_numbers(tmp, OUT)
else:
    wb.save(OUT)

try:
    shown = OUT.relative_to(ROOT)          # Path.is_relative_to 는 3.9+ 라 예외로 처리
except ValueError:
    shown = OUT
print(f"생성 완료: {shown}")
print(f"  A 용신5 · B 강약5 · C 조후2 · D 용신5 · E 설명5  ·  조합미리보기 {n-1}행")
